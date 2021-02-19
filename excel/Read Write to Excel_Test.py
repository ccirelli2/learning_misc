# -*- coding: utf-8 -*-
"""
Created on Sat Jun 24 11:19:52 2017

@author: Chris.Cirelli
"""

import openpyxl
#from openpyxl.utils.dataframe import dataframe_to_rows #use of numpy calculations


wb = openpyxl.load_workbook('new file.xlsx') #call workbook
ws = wb.get_sheet_by_name('Sheet1') #identify the name of the sheet

#==============================================================================
# #data = ws['A1'].value = 1 #create a variable that points to a cell and creates a new value.  Doesnt actually write to xl though. 
# #ws1 = ws.cell(row = 4, column = 2, value = 10) Call cell based on row/col cross section. 
# #print(wb.sheetnames) print worksheet names. 
# 
# cell_range = sheet['A1':'C2'] #use slicing to access multiple cells. 
#==============================================================================



#Ranges of rows and columns==============================================
# >>> colC = ws['C']
# >>> col_range = ws['C:D']
# >>> row10 = ws[10]
# >>> row_range = ws[5:10]
# 
#==============================================================================


##  Assign a calculation to a cell==========================================
# Delta = ws['D2'] = "=ws['C2']-ws['B2']"
# print(Delta)
# 
#==============================================================================


#==============================================================================
# df = DataFrame(ws.values)
#=========================



from openpyxl import Workbook
wb = Workbook(write_only = True)
